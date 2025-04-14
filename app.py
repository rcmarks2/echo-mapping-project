from flask import Flask, render_template, request, send_file
from geopy.geocoders import GoogleV3
from geopy.distance import geodesic
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import folium
import requests
import os
import polyline

app = Flask(__name__, template_folder="templates", static_folder="static")

# === CONFIGURATION ===
google_api_key = "AIzaSyCIPvsZMeb_NtkuElOooPCE46fB-bJEULg"
geolocator = GoogleV3(api_key=google_api_key, timeout=10)
geo_cache = {}  # Geocode result cache
route_cache = {}  # Routing distance cache

# === ROUTES ===
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/instructions")
def instructions():
    return render_template("instructions.html")

# === EV CHARGERS ===
ev_chargers = pd.concat([
    pd.read_excel(f"static/{f}.xlsx") for f in [1, 2, 3, 4, 5, 6, 7]
])
ev_chargers = ev_chargers.dropna(subset=["Latitude", "Longitude"])
ev_charger_coords = [(row["Latitude"], row["Longitude"]) for _, row in ev_chargers.iterrows()]

# === UTILITIES ===
def geocode_city_state(city, state):
    key = f"{city.strip().lower()}, {state.strip().lower()}"
    if key in geo_cache:
        return geo_cache[key]
    location = geolocator.geocode(key)
    if not location:
        raise ValueError(f"Could not geocode {city}, {state}")
    geo_cache[key] = (location.latitude, location.longitude)
    return geo_cache[key]

def get_routed_segment(start, end, return_distance=False):
    cache_key = (start, end)
    if cache_key in route_cache:
        if return_distance:
            return route_cache[cache_key]["coords"], route_cache[cache_key]["miles"]
        return route_cache[cache_key]["coords"]

    url = "https://routes.googleapis.com/directions/v2:computeRoutes"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": google_api_key,
        "X-Goog-FieldMask": "routes.polyline.encodedPolyline,routes.distanceMeters"
    }
    body = {
        "origin": {"location": {"latLng": {"latitude": start[0], "longitude": start[1]}}},
        "destination": {"location": {"latLng": {"latitude": end[0], "longitude": end[1]}}},
        "travelMode": "DRIVE"
    }

    response = requests.post(url, json=body, headers=headers)
    if response.status_code == 200:
        data = response.json()
        if "routes" in data and data["routes"]:
            encoded = data["routes"][0]["polyline"]["encodedPolyline"]
            coords = polyline.decode(encoded)
            miles = data["routes"][0]["distanceMeters"] / 1609.34
            route_cache[cache_key] = {"coords": coords, "miles": miles}
            if return_distance:
                return coords, miles
            return coords
    return [] if not return_distance else ([], 0)

def build_ev_path(start, end):
    if geodesic(start, end).miles <= 225:
        return True
    current = start
    while geodesic(current, end).miles > 225:
        next_stop = None
        for station in ev_charger_coords:
            if geodesic(current, station).miles <= 225 and geodesic(station, end).miles < geodesic(current, end).miles:
                next_stop = station
                break
        if not next_stop:
            return False
        current = next_stop
    return True

def generate_map(route_coords, used_chargers, all_chargers, label):
    m = folium.Map(location=route_coords[0], zoom_start=6, tiles="cartodbpositron")
    for coord in all_chargers:
        folium.CircleMarker(location=coord, radius=4, color="gray", fill=True, fill_opacity=0.6).add_to(m)
    for coord in used_chargers:
        folium.CircleMarker(location=coord, radius=10, color="#00cc44", fill=True, fill_opacity=1).add_to(m)
    folium.Marker(route_coords[0], popup="Start", icon=folium.DivIcon(html=f"<b>{label[0]}</b>")).add_to(m)
    folium.Marker(route_coords[-1], popup="End", icon=folium.DivIcon(html=f"<b>{label[1]}</b>")).add_to(m)
    folium.PolyLine(route_coords, color="blue", weight=4).add_to(m)
    return m._repr_html_()

@app.route("/result", methods=["POST"])
def result():
    try:
        start_city, start_state = request.form["start"].split(",")
        end_city, end_state = request.form["end"].split(",")
        mpg = float(request.form.get("mpg") or 9.0)
        trips = int(request.form["annual_trips"])
        if trips <= 0:
            raise ValueError("Annual trips must be greater than 0")
        start = geocode_city_state(start_city.strip(), start_state.strip())
        end = geocode_city_state(end_city.strip(), end_state.strip())
        diesel_coords, diesel_miles = get_routed_segment(start, end, return_distance=True)
        diesel_total = diesel_miles * trips

        hourly_rate = 75000 / 52 / 6 / 8
        diesel_hours = diesel_miles / 50
        diesel_labor_cost = hourly_rate * diesel_hours

        diesel_cost = (
            trips * (diesel_miles / mpg) * 3.59 +
            diesel_miles * (17500 / 62169) +
            diesel_total * (166000 / 750000) +
            diesel_labor_cost
        )
        diesel_emissions = (diesel_total * 1.617) / 1000
        diesel_map = generate_map(diesel_coords, [], [], (f"{start_city.strip()}, {start_state.strip()}", f"{end_city.strip()}, {end_state.strip()}"))

        ev_possible = build_ev_path(start, end)
        if ev_possible:
            ev_stops = [start]
            current = start
            while geodesic(current, end).miles > 225:
                candidates = [
                    station for station in ev_charger_coords
                    if geodesic(current, station).miles <= 225 and geodesic(station, end).miles < geodesic(current, end).miles
                ]
                if not candidates:
                    ev_possible = False
                    break
                next_stop = max(candidates, key=lambda s: geodesic(current, s).miles)
                ev_stops.append(next_stop)
                current = next_stop
            ev_stops.append(end)
            routed_coords = []
            total_ev_miles = 0
            for i in range(len(ev_stops) - 1):
                leg_coords, leg_miles = get_routed_segment(ev_stops[i], ev_stops[i + 1], return_distance=True)
                routed_coords.extend(leg_coords)
                total_ev_miles += leg_miles

            ev_hours = total_ev_miles / 50
            ev_labor_cost = hourly_rate * ev_hours

            ev_total = total_ev_miles * trips
            ev_cost = (
                (ev_total / 20.39) * 2.208 +
                total_ev_miles * (10500 / 62169) +
                ev_total * (250000 / 750000) +
                ev_labor_cost
            )
            ev_emissions = (ev_total * 0.2102) / 1000
            ev_map = generate_map(routed_coords, ev_stops[1:-1], ev_charger_coords, (f"{start_city.strip()}, {start_state.strip()}", f"{end_city.strip()}, {end_state.strip()}"))
        else:
            ev_map = None
            ev_total = ev_cost = ev_emissions = None

        return render_template("result.html",
            diesel_miles=round(diesel_miles, 1),
            annual_trips=trips,
            diesel_annual_miles=round(diesel_total, 1),
            diesel_total_cost=f"{diesel_cost:,.2f}",
            diesel_emissions=round(diesel_emissions, 2),
            ev_unavailable=not ev_possible,
            ev_miles=round(total_ev_miles, 1) if ev_possible else None,
            ev_annual_miles=round(ev_total, 1) if ev_possible else None,
            ev_total_cost=f"{ev_cost:,.2f}" if ev_possible else None,
            ev_emissions=round(ev_emissions, 2) if ev_possible else None,
            diesel_map=diesel_map,
            ev_map=ev_map
        )
    except Exception as e:
        return f"<h3>Error in single route: {e}</h3>"
@app.route("/batch-result", methods=["POST"])
def batch_result():
    try:
        uploaded_file = request.files['excel']
        if uploaded_file.filename == '':
            return '<h3>No file selected</h3>'

        df = pd.read_excel(uploaded_file)
        wb = load_workbook('static/fullbatchresult.xlsx')
        ws = wb.active

        def process_row(i, row):
            try:
                start_city = str(row["Start City"]).strip()
                start_state = str(row["Start State"]).strip()
                dest_city = str(row["Destination City"]).strip()
                dest_state = str(row["Destination State"]).strip()
                trips = max(int(row.get("Annual Trips (Minimum 1)", 1)), 1)
                mpg = float(row.get("MPG (Optional Will Default to 9)", 9) or 9)

                start = geocode_city_state(start_city, start_state)
                end = geocode_city_state(dest_city, dest_state)
                _, diesel_miles = get_routed_segment(start, end, return_distance=True)

                ev_possible = "No"
                ev_miles = "N/A"

                if build_ev_path(start, end):
                    ev_stops = [start]
                    current = start
                    feasible = True
                    while geodesic(current, end).miles > 225:
                        candidates = [
                            station for station in ev_charger_coords
                            if geodesic(current, station).miles <= 225 and geodesic(station, end).miles < geodesic(current, end).miles
                        ]
                        if not candidates:
                            feasible = False
                            break
                        next_stop = max(candidates, key=lambda s: geodesic(current, s).miles)
                        ev_stops.append(next_stop)
                        current = next_stop
                    if feasible:
                        ev_stops.append(end)
                        total_ev_miles = 0
                        for j in range(len(ev_stops) - 1):
                            _, leg_miles = get_routed_segment(ev_stops[j], ev_stops[j + 1], return_distance=True)
                            total_ev_miles += leg_miles
                        ev_possible = "Yes"
                        ev_miles = round(total_ev_miles, 1)

                ws.cell(row=i + 3, column=1).value = start_city
                ws.cell(row=i + 3, column=2).value = start_state
                ws.cell(row=i + 3, column=3).value = dest_city
                ws.cell(row=i + 3, column=4).value = dest_state
                ws.cell(row=i + 3, column=5).value = round(diesel_miles, 1)
                ws.cell(row=i + 3, column=6).value = trips
                ws.cell(row=i + 3, column=7).value = ev_possible
                ws.cell(row=i + 3, column=8).value = ev_miles
            except Exception as err:
                ws.cell(row=i + 3, column=1).value = f"Error: {str(err)}"

        # 🚀 Process rows in parallel using threads
        with ThreadPoolExecutor(max_workers=20) as executor:
            for i in range(len(df)):
                executor.submit(process_row, i, df.iloc[i])

        wb.save('static/fullbatchresult.xlsx')
        return render_template('batch_result.html', excel_download='/download-batch-excel', txt_download='/download-formulas')
    except Exception as e:
        return f'<h3>Error in batch processing: {e}</h3>'

@app.route("/download-batch-excel")
def download_batch_excel():
    return send_file("static/fullbatchresult.xlsx", as_attachment=True)

@app.route("/download-formulas")
def download_formulas():
    return send_file("static/formulas.txt", as_attachment=True)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
